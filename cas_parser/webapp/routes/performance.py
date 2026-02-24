import json
import urllib.request
import urllib.parse
from datetime import datetime, timedelta
from flask import Blueprint, jsonify, request, send_file
from flask import current_app
from cas_parser.webapp import data as db
from cas_parser.webapp.xirr import build_cashflows_for_folio, xirr, _parse_date
from cas_parser.webapp.routes import DecimalEncoder
from cas_parser.webapp.auth import admin_required, check_investor_access, get_investor_id_for_folio

performance_bp = Blueprint('performance', __name__)


@performance_bp.route('/api/investors/<int:investor_id>/performance', methods=['GET'])
def api_get_performance(investor_id):
    """Get complete performance data with metrics."""
    check_investor_access(investor_id)
    category = request.args.get('category', '')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    # Get user-added extra benchmarks
    extra_benchmarks = db.get_benchmarks_by_investor(investor_id)

    from cas_parser.webapp.benchmarking import get_performance_data
    try:
        result = get_performance_data(
            investor_id, category or None, start_date, end_date,
            extra_benchmarks=extra_benchmarks
        )
        return jsonify(result)
    except Exception as e:
        current_app.logger.error(f"Performance calculation error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@performance_bp.route('/api/investors/<int:investor_id>/performance/export', methods=['GET'])
def api_export_performance(investor_id):
    """Export XIRR cashflows as Excel file for validation.

    Sheet 1: Portfolio cashflows.
    Sheet 2+: One sheet per user-added benchmark.
    """
    check_investor_access(investor_id)
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font

    category = request.args.get('category', '')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    # Get user-added benchmarks
    user_benchmarks = db.get_benchmarks_by_investor(investor_id)

    from cas_parser.webapp.benchmarking import get_cashflows_for_export
    try:
        data = get_cashflows_for_export(
            investor_id, category or None, start_date, end_date,
            benchmarks=user_benchmarks
        )
    except Exception as e:
        current_app.logger.error(f"Export error: {e}")
        return jsonify({'error': str(e)}), 500

    wb = Workbook()
    bold = Font(bold=True)
    date_fmt = 'YYYY-MM-DD'
    num_fmt = '#,##0.00'

    # --- Sheet 1: Portfolio Cashflows ---
    ws1 = wb.active
    ws1.title = 'Portfolio Cashflows'
    headers1 = ['Date', 'Description', 'Amount']
    for col, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=col, value=h).font = bold

    for i, row in enumerate(data['portfolio_cashflows'], 2):
        try:
            ws1.cell(row=i, column=1, value=datetime.strptime(
                row['date'], '%Y-%m-%d'
            )).number_format = date_fmt
        except (ValueError, TypeError):
            ws1.cell(row=i, column=1, value=row['date'])
        ws1.cell(row=i, column=2, value=row['description'])
        ws1.cell(row=i, column=3, value=row['amount']).number_format = num_fmt

    xirr_row = len(data['portfolio_cashflows']) + 3
    ws1.cell(row=xirr_row, column=2, value='XIRR').font = bold
    xirr_val = data.get('portfolio_xirr')
    if xirr_val is not None:
        ws1.cell(
            row=xirr_row, column=3, value=xirr_val / 100
        ).number_format = '0.00%'
    else:
        ws1.cell(row=xirr_row, column=3, value='N/A')

    _auto_width(ws1, 3)

    # --- Benchmark sheets: one per user-added benchmark ---
    for bm in data.get('benchmarks', []):
        bm_name = bm.get('name', 'Benchmark')
        # Excel sheet names: max 31 chars, no : \ / ? * [ ]
        safe = bm_name.translate(
            str.maketrans(':\\/?*[]', '       ')
        ).strip()
        sheet_title = safe[:31]

        ws = wb.create_sheet(title=sheet_title)
        headers = ['Date', 'Description', 'Amount', 'Benchmark NAV',
                   'Units', 'Cumulative Units']
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h).font = bold

        for i, row in enumerate(bm.get('cashflows', []), 2):
            try:
                ws.cell(row=i, column=1, value=datetime.strptime(
                    row['date'], '%Y-%m-%d'
                )).number_format = date_fmt
            except (ValueError, TypeError):
                ws.cell(row=i, column=1, value=row['date'])
            ws.cell(row=i, column=2, value=row['description'])
            ws.cell(
                row=i, column=3, value=row['amount']
            ).number_format = num_fmt
            ws.cell(
                row=i, column=4, value=row.get('nav', '')
            ).number_format = '#,##0.0000'
            units = row.get('units', '')
            if units == 0 and row['description'] == 'Terminal Value':
                ws.cell(row=i, column=5, value='\u2014')
            else:
                ws.cell(
                    row=i, column=5, value=units
                ).number_format = '#,##0.0000'
            ws.cell(
                row=i, column=6, value=row.get('cumulative_units', '')
            ).number_format = '#,##0.0000'

        xirr_r = len(bm.get('cashflows', [])) + 3
        ws.cell(row=xirr_r, column=2, value='XIRR').font = bold
        bm_xirr = bm.get('xirr')
        if bm_xirr is not None:
            ws.cell(
                row=xirr_r, column=3, value=bm_xirr / 100
            ).number_format = '0.00%'
        else:
            ws.cell(row=xirr_r, column=3, value='N/A')

        _auto_width(ws, 6)

    # Save and send
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    cat_label = data.get('category', 'all')
    filename = f'xirr_cashflows_{cat_label}.xlsx'

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument'
                 '.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


def _auto_width(ws, num_cols):
    """Auto-size column widths for an openpyxl worksheet."""
    for col in range(1, num_cols + 1):
        max_len = max(
            (len(str(ws.cell(row=r, column=col).value or ''))
             for r in range(1, ws.max_row + 1)),
            default=10
        )
        ws.column_dimensions[
            ws.cell(row=1, column=col).column_letter
        ].width = min(max_len + 4, 30)


@performance_bp.route('/api/investors/<int:investor_id>/performance/returns', methods=['GET'])
def api_multi_period_returns(investor_id):
    """Get portfolio XIRR vs benchmark CAGR for all standard periods."""
    check_investor_access(investor_id)
    category = request.args.get('category', '')

    extra_benchmarks = db.get_benchmarks_by_investor(investor_id)

    from cas_parser.webapp.benchmarking import get_multi_period_returns
    try:
        result = get_multi_period_returns(
            investor_id, category or None,
            extra_benchmarks=extra_benchmarks
        )
        return jsonify(result)
    except Exception as e:
        current_app.logger.error(f"Multi-period returns error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@performance_bp.route('/api/investors/<int:investor_id>/benchmarks', methods=['GET'])
def api_get_benchmarks(investor_id):
    """Get all saved benchmarks for an investor."""
    check_investor_access(investor_id)
    benchmarks = db.get_benchmarks_by_investor(investor_id)
    return jsonify(benchmarks)


@performance_bp.route('/api/investors/<int:investor_id>/benchmarks', methods=['POST'])
def api_add_benchmark(investor_id):
    """Add a benchmark for an investor and fetch its NAV data."""
    check_investor_access(investor_id)
    data = request.json
    if not data or 'scheme_code' not in data or 'scheme_name' not in data:
        return jsonify({'error': 'Missing scheme_code or scheme_name'}), 400

    benchmark_id = db.add_benchmark(
        investor_id,
        int(data['scheme_code']),
        data['scheme_name'],
        data.get('fund_house')
    )

    # Fetch and cache NAV data from MFAPI
    try:
        _fetch_and_cache_benchmark(int(data['scheme_code']))
    except Exception as e:
        current_app.logger.warning(f"Failed to fetch benchmark data for {data['scheme_code']}: {e}")

    return jsonify({'id': benchmark_id, 'scheme_code': data['scheme_code'], 'scheme_name': data['scheme_name']})


@performance_bp.route('/api/investors/<int:investor_id>/benchmarks/<int:benchmark_id>', methods=['DELETE'])
def api_delete_benchmark(investor_id, benchmark_id):
    """Delete a benchmark."""
    check_investor_access(investor_id)
    deleted = db.delete_benchmark(investor_id, benchmark_id)
    if not deleted:
        return jsonify({'error': 'Benchmark not found'}), 404
    return jsonify({'success': True})


@performance_bp.route('/api/mfapi/search', methods=['GET'])
def api_mfapi_search():
    """Proxy search to MFAPI.in."""
    q = request.args.get('q', '').strip()
    if not q:
        return jsonify([])
    try:
        url = 'https://api.mfapi.in/mf/search?q=' + urllib.parse.quote(q)
        req = urllib.request.Request(url, headers={'User-Agent': 'FamFolioz/1.0'})
        with urllib.request.urlopen(req, timeout=10) as resp:
            results = json.loads(resp.read().decode('utf-8'))
        return jsonify(results)
    except Exception as e:
        current_app.logger.error(f"MFAPI search error: {e}")
        return jsonify({'error': 'Failed to search MFAPI'}), 502


@performance_bp.route('/api/benchmarks/<int:scheme_code>/data', methods=['GET'])
def api_get_benchmark_data(scheme_code):
    """Get benchmark NAV data, fetching from MFAPI if stale or missing."""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    # Check cache freshness
    latest = db.get_benchmark_data_latest_date(scheme_code)
    needs_refresh = False
    if not latest:
        needs_refresh = True
    else:
        latest_dt = datetime.strptime(latest, '%Y-%m-%d')
        if (datetime.now() - latest_dt).days > 2:
            needs_refresh = True

    if needs_refresh:
        try:
            _fetch_and_cache_benchmark(scheme_code)
        except Exception as e:
            current_app.logger.warning(f"Failed to refresh benchmark {scheme_code}: {e}")

    data = db.get_benchmark_data(scheme_code, start_date, end_date)
    return jsonify(data)


def _fetch_and_cache_benchmark(scheme_code: int):
    """Fetch NAV data from MFAPI.in and cache in benchmark_data table."""
    url = f'https://api.mfapi.in/mf/{scheme_code}'
    req = urllib.request.Request(url, headers={'User-Agent': 'FamFolioz/1.0'})
    with urllib.request.urlopen(req, timeout=30) as resp:
        result = json.loads(resp.read().decode('utf-8'))

    nav_data = result.get('data', [])
    if not nav_data:
        return

    rows = []
    for entry in nav_data:
        try:
            # MFAPI dates are DD-MM-YYYY
            dt = datetime.strptime(entry['date'], '%d-%m-%Y')
            rows.append({
                'data_date': dt.strftime('%Y-%m-%d'),
                'nav': float(entry['nav'])
            })
        except (ValueError, KeyError):
            continue

    db.upsert_benchmark_data(scheme_code, rows)


@performance_bp.route('/api/nav/refresh', methods=['POST'])
@admin_required
def api_refresh_nav():
    """Refresh NAV data from AMFI for mapped funds and take portfolio snapshots."""
    try:
        result = db.fetch_and_update_nav()

        # Also take portfolio snapshots for all investors
        if result.get('success'):
            snapshot_result = db.take_all_portfolio_snapshots()
            result['snapshots'] = snapshot_result

        return jsonify(result)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@performance_bp.route('/api/nav/history/<isin>', methods=['GET'])
def api_get_nav_history(isin):
    """Get historical NAV for a scheme."""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    history = db.get_nav_history(isin, start_date, end_date)
    return jsonify(history)


@performance_bp.route('/api/nav/history-dates', methods=['GET'])
def api_get_nav_history_dates():
    """Get all dates with NAV history."""
    dates = db.get_nav_history_dates()
    return jsonify(dates)


@performance_bp.route('/api/nav/status', methods=['GET'])
def api_nav_status():
    """Get NAV update status."""
    last_update = db.get_last_nav_update()
    stats = db.get_mutual_fund_stats()
    return jsonify({
        'last_update': last_update,
        'stats': stats
    })


@performance_bp.route('/api/folios/<int:folio_id>/xirr', methods=['GET'])
def api_get_folio_xirr(folio_id):
    """Get XIRR for a specific folio."""
    inv_id = get_investor_id_for_folio(folio_id)
    if inv_id:
        check_investor_access(inv_id)
    data = db.get_xirr_data_for_folio(folio_id)
    cashflows = build_cashflows_for_folio(
        data['transactions'], data['current_value']
    )
    xirr_val = xirr(cashflows)
    return jsonify({
        'folio_id': folio_id,
        'xirr': round(xirr_val * 100, 2) if xirr_val is not None else None,
        'current_value': data['current_value'],
        'cashflow_count': len(cashflows),
    })


@performance_bp.route('/api/investors/<int:investor_id>/xirr', methods=['GET'])
def api_get_investor_xirr(investor_id):
    """Get XIRR for all folios of an investor + portfolio-level and per-ISIN XIRR."""
    check_investor_access(investor_id)
    folio_data_list = db.get_xirr_data_for_investor(investor_id)
    all_cashflows = []
    folios = []
    isin_cashflows = {}  # isin -> list of cashflows

    for data in folio_data_list:
        cashflows = build_cashflows_for_folio(
            data['transactions'], data['current_value']
        )
        xirr_val = xirr(cashflows)
        folio_isin = data.get('isin')
        folios.append({
            'folio_id': data['folio_id'],
            'scheme_name': data['scheme_name'],
            'folio_number': data['folio_number'],
            'isin': folio_isin,
            'xirr': round(xirr_val * 100, 2) if xirr_val is not None else None,
            'current_value': data['current_value'],
            'cashflow_count': len(cashflows),
        })
        # Only include folios with valid XIRR in portfolio calculation
        # (folios with bad data that can't compute XIRR would corrupt portfolio XIRR)
        if xirr_val is not None:
            all_cashflows.extend(cashflows)
            # Accumulate per-ISIN cashflows
            if folio_isin:
                if folio_isin not in isin_cashflows:
                    isin_cashflows[folio_isin] = []
                isin_cashflows[folio_isin].extend(cashflows)

    # Include manual assets (PPF/EPF etc.) in portfolio XIRR
    manual_asset_data = db.get_manual_asset_xirr_data(investor_id)
    for asset_data in manual_asset_data:
        # Convert string dates to date objects for xirr()
        asset_cashflows = []
        for date_str, amount in asset_data['cashflows']:
            d = _parse_date(date_str)
            if d is not None:
                asset_cashflows.append((d, amount))

        asset_xirr_val = xirr(asset_cashflows) if len(asset_cashflows) >= 2 else None
        folios.append({
            'folio_id': f"manual_{asset_data['asset_id']}",
            'scheme_name': asset_data['asset_name'],
            'folio_number': asset_data['asset_type'].upper(),
            'isin': None,
            'xirr': round(asset_xirr_val * 100, 2) if asset_xirr_val is not None else None,
            'current_value': asset_data['current_value'],
            'cashflow_count': len(asset_cashflows),
        })
        if asset_xirr_val is not None:
            all_cashflows.extend(asset_cashflows)

    # Compute per-ISIN aggregated XIRR
    isin_xirr = {}
    for isin, cfs in isin_cashflows.items():
        xirr_val = xirr(cfs)
        isin_xirr[isin] = round(xirr_val * 100, 2) if xirr_val is not None else None

    portfolio_xirr = xirr(all_cashflows)
    return jsonify({
        'portfolio_xirr': round(portfolio_xirr * 100, 2) if portfolio_xirr is not None else None,
        'folios': folios,
        'isin_xirr': isin_xirr,
    })


@performance_bp.route('/api/investors/<int:investor_id>/portfolio-history', methods=['GET'])
def api_get_portfolio_history(investor_id):
    """Get historical portfolio valuation."""
    check_investor_access(investor_id)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    history = db.get_portfolio_history(investor_id, start_date, end_date)
    return jsonify(history)


@performance_bp.route('/api/investors/<int:investor_id>/valuation/<valuation_date>', methods=['GET'])
def api_get_portfolio_valuation(investor_id, valuation_date):
    """Get portfolio valuation on a specific date."""
    check_investor_access(investor_id)
    valuation = db.get_portfolio_valuation_on_date(investor_id, valuation_date)
    return jsonify(valuation)


@performance_bp.route('/api/investors/<int:investor_id>/snapshot', methods=['POST'])
def api_take_portfolio_snapshot(investor_id):
    """Take a portfolio snapshot for an investor."""
    check_investor_access(investor_id)
    data = request.json or {}
    snapshot_date = data.get('snapshot_date')
    result = db.take_portfolio_snapshot(investor_id, snapshot_date)
    return jsonify(result)
